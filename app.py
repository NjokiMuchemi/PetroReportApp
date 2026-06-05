import io
import re
from pathlib import Path
from datetime import datetime, date
from typing import Dict, List, Tuple, Optional

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PRODUCT_MAP = {
    "PMS": "PMS", "ULP": "PMS", "UNLEADED": "PMS", "RMS": "PMS", "PETROL": "PMS", "PMS.": "PMS",
    "AGO": "AGO", "DIESEL": "AGO", "DSL": "AGO", "AUTOMOTIVE GAS OIL": "AGO",
    "IK": "IK", "BIK": "IK", "KEROSENE": "IK", "KERO": "IK", "ILLUMINATING KEROSENE": "IK",
}
PRODUCT_ORDER = ["PMS", "AGO", "IK"]


def clean_text(x):
    if x is None:
        return ""
    return str(x).strip()


def norm_text(x):
    return re.sub(r"\s+", " ", clean_text(x).upper())


def norm_station(x):
    s = clean_text(x)
    s = re.sub(r"\s+", " ", s)
    s = s.replace("AVENUE", "").replace("DEALER", "")
    return s.strip().title()


def is_total_station(station):
    """Remove workbook total columns/rows that are not petrol stations."""
    s = norm_text(station)
    if not s:
        return True
    total_words = [
        "TOTAL", "TOTALS", "TOTALS TODATE", "TOTALS TO DATE",
        "TOTALS LOSSES TODATE", "TOTAL LOSSES TODATE",
        "ZONE TOTAL", "GRAND TOTAL", "NETWORK TOTAL",
        "CUMM", "CUMULATIVE"
    ]
    return any(word in s for word in total_words)


def norm_product(x):
    key = norm_text(x).replace(".", "")
    return PRODUCT_MAP.get(key, key)


def is_date_like(x):
    return isinstance(x, (datetime, date)) or (isinstance(x, str) and bool(re.search(r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}", x)))


def to_date(x):
    if isinstance(x, datetime):
        return x.date()
    if isinstance(x, date):
        return x
    try:
        return pd.to_datetime(x).date()
    except Exception:
        return None


def workbook_from_upload(uploaded_file):
    return load_workbook(io.BytesIO(uploaded_file.getvalue()), data_only=True, read_only=False, keep_links=False)


def find_sheet_name(wb, wanted: str):
    wanted_norm = norm_text(wanted)
    for s in wb.sheetnames:
        if norm_text(s) == wanted_norm:
            return s
    for s in wb.sheetnames:
        if wanted_norm in norm_text(s):
            return s
    return None


def find_product_header(ws):
    for row in range(1, min(ws.max_row, 15) + 1):
        vals = [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
        for col, v in enumerate(vals, start=1):
            if norm_text(v) in ["PRODUCT", "PDT"]:
                return row, col
    raise ValueError(f"Could not find PRODUCT header in sheet {ws.title}")


def extract_product_summary(wb, zone_name: str) -> pd.DataFrame:
    sname = find_sheet_name(wb, "PER PRODUCT SUMMARY")
    if not sname:
        return pd.DataFrame(columns=["Zone", "Date", "Station", "Product", "Receipts", "Sales"])
    ws = wb[sname]
    header_row, product_col = find_product_header(ws)
    station_row = header_row - 1
    date_col = product_col - 1
    records = []

    station_by_col = {}
    for col in range(product_col + 1, ws.max_column + 1):
        metric = norm_text(ws.cell(header_row, col).value)
        if "METERED" in metric or "SALES" == metric:
            station = ws.cell(station_row, col - 1).value or ws.cell(station_row, col).value
            receipt_col = col - 1
            station_by_col[col] = (norm_station(station), receipt_col)

    for row in range(header_row + 1, ws.max_row + 1):
        dt = to_date(ws.cell(row, date_col).value)
        product = norm_product(ws.cell(row, product_col).value)
        if not dt or product not in PRODUCT_ORDER:
            continue
        for sales_col, (station, receipt_col) in station_by_col.items():
            if not station or is_total_station(station):
                continue
            receipts = ws.cell(row, receipt_col).value or 0
            sales = ws.cell(row, sales_col).value or 0
            if receipts == 0 and sales == 0:
                # keep rows only when some value exists; blank station/day rows add noise
                continue
            records.append({
                "Zone": zone_name, "Date": dt, "Station": station, "Product": product,
                "Receipts": float(receipts or 0), "Sales": float(sales or 0)
            })
    return pd.DataFrame(records)


def extract_product_loss(wb, zone_name: str) -> pd.DataFrame:
    sname = find_sheet_name(wb, "PER PDT LOSS")
    if not sname:
        return pd.DataFrame(columns=["Zone", "Date", "Station", "Product", "Loss"])
    ws = wb[sname]
    header_row, product_col = find_product_header(ws)
    station_row = header_row - 1
    date_col = product_col - 1
    records = []

    loss_cols = {}
    for col in range(product_col + 1, ws.max_column + 1):
        metric = norm_text(ws.cell(header_row, col).value)
        if metric == "LOSS":
            station = ws.cell(station_row, col).value or ws.cell(station_row, col - 1).value
            loss_cols[col] = norm_station(station)

    for row in range(header_row + 1, ws.max_row + 1):
        dt = to_date(ws.cell(row, date_col).value)
        product = norm_product(ws.cell(row, product_col).value)
        if not dt or product not in PRODUCT_ORDER:
            continue
        for col, station in loss_cols.items():
            if not station or is_total_station(station):
                continue
            loss = ws.cell(row, col).value or 0
            if isinstance(loss, str):
                continue
            if loss == 0:
                continue
            records.append({"Zone": zone_name, "Date": dt, "Station": station, "Product": product, "Loss": float(loss or 0)})
    return pd.DataFrame(records)


def extract_banking(wb, zone_name: str) -> pd.DataFrame:
    sname = find_sheet_name(wb, "BANKING FIGURES") or find_sheet_name(wb, "BANKING")
    if not sname:
        return pd.DataFrame(columns=["Zone", "Date", "Station", "Banking"])
    ws = wb[sname]
    header_row = None
    date_col = None
    for row in range(1, min(ws.max_row, 15) + 1):
        for col in range(1, ws.max_column + 1):
            if norm_text(ws.cell(row, col).value) == "DATE":
                header_row, date_col = row, col
                break
        if header_row:
            break
    if not header_row:
        return pd.DataFrame(columns=["Zone", "Date", "Station", "Banking"])

    station_cols = []
    stop_words = ["TOTAL", "CUMM", "CUM", "BALANCE"]
    for col in range(date_col + 1, ws.max_column + 1):
        name = norm_text(ws.cell(header_row, col).value)
        if not name:
            continue
        if any(w in name for w in stop_words):
            continue
        station = norm_station(name)
        if is_total_station(station):
            continue
        station_cols.append((col, station))

    records = []
    for row in range(header_row + 1, ws.max_row + 1):
        dt = to_date(ws.cell(row, date_col).value)
        if not dt:
            continue
        for col, station in station_cols:
            val = ws.cell(row, col).value or 0
            if isinstance(val, str):
                continue
            if val == 0:
                continue
            records.append({"Zone": zone_name, "Date": dt, "Station": station, "Banking": float(val or 0)})
    return pd.DataFrame(records)


def extract_targets(uploaded_file) -> pd.DataFrame:
    if uploaded_file is None:
        return pd.DataFrame(columns=["Station", "Target"])
    wb = workbook_from_upload(uploaded_file)
    ws = wb[wb.sheetnames[0]]
    header_row = None
    headers = {}
    for row in range(1, min(ws.max_row, 20) + 1):
        vals = [norm_text(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
        if "STATION" in vals and ("MONTH TARGET" in vals or "TARGET" in vals):
            header_row = row
            headers = {vals[col-1]: col for col in range(1, ws.max_column + 1)}
            break
    if not header_row:
        return pd.DataFrame(columns=["Station", "Target"])
    station_col = headers.get("STATION")
    target_col = headers.get("MONTH TARGET") or headers.get("TARGET")
    rows = []
    for row in range(header_row + 1, ws.max_row + 1):
        stn = norm_station(ws.cell(row, station_col).value)
        target = ws.cell(row, target_col).value
        if stn and isinstance(target, (int, float)):
            rows.append({"Station": stn, "Target": float(target)})
    return pd.DataFrame(rows).drop_duplicates("Station")



TARGETS_FILE = Path(__file__).parent / "data" / "monthly_targets.csv"
STATION_MASTER_FILE = Path(__file__).parent / "data" / "station_master.csv"

DEFAULT_STATIONS = [
    ("COAST", "Bombolulu"), ("COAST", "Buxton"), ("COAST", "Diani"), ("COAST", "Ferry"),
    ("COAST", "Kaloleni"), ("COAST", "Lamu"), ("COAST", "Langalanga"), ("COAST", "Likoni"),
    ("COAST", "Malindi"), ("COAST", "Mishomoroni"), ("COAST", "Moi"), ("COAST", "Sabasaba"),
    ("COAST", "Shelly Beach"), ("COAST", "Tononoka"), ("COAST", "Utange"), ("COAST", "Voi"),
    ("NAIROBI", "Bypass"), ("NAIROBI", "Matuu"), ("NAIROBI", "Meru"), ("NAIROBI", "Ngara"),
    ("NAIROBI", "Othaya"), ("NAIROBI", "Pangani"), ("NAIROBI", "Transhighway"), ("NAIROBI", "Waiyaki"),
    ("WESTERN", "Busia"), ("WESTERN", "Cheramei"), ("WESTERN", "Huruma"), ("WESTERN", "Lolwe"),
    ("WESTERN", "Migori"), ("WESTERN", "Mumia"), ("WESTERN", "Siaya"), ("WESTERN", "Webuye"),
]

DEFAULT_ALIASES = {
    "Waiyaki Way": "Waiyaki",
    "Moi Avenue": "Moi",
    "Moi Av": "Moi",
    "Moi Ave": "Moi",
    "Shelly": "Shelly Beach",
}


def ensure_targets_store():
    TARGETS_FILE.parent.mkdir(parents=True, exist_ok=True)


def load_saved_targets() -> pd.DataFrame:
    ensure_targets_store()
    if TARGETS_FILE.exists():
        try:
            df = pd.read_csv(TARGETS_FILE)
            if "Station" not in df.columns or "Target" not in df.columns:
                return pd.DataFrame(columns=["Station", "Target"])
            df = df[["Station", "Target"]].copy()
            df["Station"] = df["Station"].apply(norm_station)
            df["Target"] = pd.to_numeric(df["Target"], errors="coerce").fillna(0.0)
            df = df[df["Station"].astype(str).str.strip() != ""]
            df = df[~df["Station"].apply(is_total_station)]
            return df.drop_duplicates("Station").sort_values("Station").reset_index(drop=True)
        except Exception:
            return pd.DataFrame(columns=["Station", "Target"])
    return pd.DataFrame(columns=["Station", "Target"])


def save_targets(df: pd.DataFrame):
    ensure_targets_store()
    if df is None or df.empty:
        clean = pd.DataFrame(columns=["Station", "Target"])
    else:
        clean = df[["Station", "Target"]].copy()
        clean["Station"] = clean["Station"].apply(norm_station)
        clean["Target"] = pd.to_numeric(clean["Target"], errors="coerce").fillna(0.0)
        clean = clean[clean["Station"].astype(str).str.strip() != ""]
        clean = clean[~clean["Station"].apply(is_total_station)]
        clean = clean.drop_duplicates("Station").sort_values("Station").reset_index(drop=True)
    clean.to_csv(TARGETS_FILE, index=False)
    return clean


def ensure_station_master_store():
    STATION_MASTER_FILE.parent.mkdir(parents=True, exist_ok=True)
    if not STATION_MASTER_FILE.exists():
        rows = []
        for zone, station in DEFAULT_STATIONS:
            rows.append({"Zone": zone, "Standard Station": station, "Alias": station})
        for alias, standard in DEFAULT_ALIASES.items():
            zone = next((z for z, st in DEFAULT_STATIONS if st == standard), "")
            rows.append({"Zone": zone, "Standard Station": standard, "Alias": alias})
        pd.DataFrame(rows).to_csv(STATION_MASTER_FILE, index=False)


def load_station_master() -> pd.DataFrame:
    ensure_station_master_store()
    try:
        df = pd.read_csv(STATION_MASTER_FILE)
    except Exception:
        df = pd.DataFrame(columns=["Zone", "Standard Station", "Alias"])
    for col in ["Zone", "Standard Station", "Alias"]:
        if col not in df.columns:
            df[col] = ""
    df = df[["Zone", "Standard Station", "Alias"]].copy()
    df["Zone"] = df["Zone"].apply(lambda x: norm_text(x))
    df["Standard Station"] = df["Standard Station"].apply(norm_station)
    df["Alias"] = df["Alias"].apply(norm_station)
    df = df[(df["Standard Station"] != "") & (df["Alias"] != "")]
    df = df[~df["Standard Station"].apply(is_total_station)]
    return df.drop_duplicates(["Zone", "Alias"]).sort_values(["Zone", "Standard Station", "Alias"]).reset_index(drop=True)


def save_station_master(df: pd.DataFrame) -> pd.DataFrame:
    ensure_station_master_store()
    if df is None or df.empty:
        clean = pd.DataFrame(columns=["Zone", "Standard Station", "Alias"])
    else:
        clean = df.copy()
        for col in ["Zone", "Standard Station", "Alias"]:
            if col not in clean.columns:
                clean[col] = ""
        clean = clean[["Zone", "Standard Station", "Alias"]]
        clean["Zone"] = clean["Zone"].apply(lambda x: norm_text(x))
        clean["Standard Station"] = clean["Standard Station"].apply(norm_station)
        clean["Alias"] = clean["Alias"].apply(norm_station)
        clean = clean[(clean["Standard Station"] != "") & (clean["Alias"] != "")]
        clean = clean[~clean["Standard Station"].apply(is_total_station)]
        clean = clean.drop_duplicates(["Zone", "Alias"]).sort_values(["Zone", "Standard Station", "Alias"]).reset_index(drop=True)
    clean.to_csv(STATION_MASTER_FILE, index=False)
    return clean


def apply_station_master_to_df(df: pd.DataFrame, station_master: pd.DataFrame, source_name: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    if df is None or df.empty or "Station" not in df.columns:
        return df, pd.DataFrame(columns=["Source", "Zone", "Unmapped Station", "Rows"])

    master = station_master.copy() if station_master is not None else pd.DataFrame(columns=["Zone", "Standard Station", "Alias"])
    if master.empty:
        mapped = df.copy()
        mapped["Original Station"] = mapped["Station"]
        return mapped.iloc[0:0], pd.DataFrame(columns=["Source", "Zone", "Unmapped Station", "Rows"])

    # Map by approved Zone + Alias only. No guessing.
    lookup = {(norm_text(r["Zone"]), norm_station(r["Alias"])): norm_station(r["Standard Station"]) for _, r in master.iterrows()}
    out = df.copy()
    out["Original Station"] = out["Station"].apply(norm_station)
    mapped_station = []
    is_mapped = []
    for _, row in out.iterrows():
        key = (norm_text(row.get("Zone", "")), norm_station(row.get("Station", "")))
        standard = lookup.get(key)
        mapped_station.append(standard or norm_station(row.get("Station", "")))
        is_mapped.append(bool(standard))
    out["Station"] = mapped_station
    out["_StationMapped"] = is_mapped

    unmapped = out[~out["_StationMapped"]].groupby(["Zone", "Original Station"], as_index=False).size()
    if not unmapped.empty:
        unmapped = unmapped.rename(columns={"Original Station": "Unmapped Station", "size": "Rows"})
        unmapped.insert(0, "Source", source_name)
    else:
        unmapped = pd.DataFrame(columns=["Source", "Zone", "Unmapped Station", "Rows"])

    out = out[out["_StationMapped"]].drop(columns=["_StationMapped"])
    return out, unmapped


def build_reports(sales_df, loss_df, banking_df, targets_df):
    # MTD Sales
    sales_pivot = pd.pivot_table(sales_df, values="Sales", index=["Zone", "Station"], columns="Product", aggfunc="sum", fill_value=0).reset_index()
    for p in PRODUCT_ORDER:
        if p not in sales_pivot.columns:
            sales_pivot[p] = 0.0
    sales_pivot = sales_pivot[["Zone", "Station"] + PRODUCT_ORDER]
    sales_pivot["Total Sales"] = sales_pivot[PRODUCT_ORDER].sum(axis=1)

    days = sales_df.groupby(["Zone", "Station"])["Date"].nunique().reset_index(name="Days Counted")
    perf = sales_pivot.merge(days, on=["Zone", "Station"], how="left")
    perf["Average Sales/Day"] = perf.apply(lambda r: r["Total Sales"] / r["Days Counted"] if r["Days Counted"] else 0, axis=1)

    # Loss
    loss_pivot = pd.pivot_table(loss_df, values="Loss", index=["Zone", "Station"], columns="Product", aggfunc="sum", fill_value=0).reset_index()
    for p in PRODUCT_ORDER:
        if p not in loss_pivot.columns:
            loss_pivot[p] = 0.0
    loss_pivot = loss_pivot.rename(columns={"PMS": "Loss PMS", "AGO": "Loss AGO", "IK": "Loss IK"})
    perf = perf.merge(loss_pivot, on=["Zone", "Station"], how="left")
    for c in ["Loss PMS", "Loss AGO", "Loss IK"]:
        if c not in perf.columns:
            perf[c] = 0.0
        perf[c] = perf[c].fillna(0.0)
    perf["Total Loss"] = perf[["Loss PMS", "Loss AGO", "Loss IK"]].sum(axis=1)
    perf["% Loss to Sales"] = perf.apply(lambda r: r["Total Loss"] / r["Total Sales"] if r["Total Sales"] else 0, axis=1)
    perf["% PMS Loss"] = perf.apply(lambda r: r["Loss PMS"] / r["PMS"] if r["PMS"] else 0, axis=1)
    perf["% AGO Loss"] = perf.apply(lambda r: r["Loss AGO"] / r["AGO"] if r["AGO"] else 0, axis=1)
    perf["% IK Loss"] = perf.apply(lambda r: r["Loss IK"] / r["IK"] if r["IK"] else 0, axis=1)

    if not targets_df.empty:
        perf = perf.merge(targets_df, on="Station", how="left")
    else:
        perf["Target"] = 0.0
    perf["Target"] = perf["Target"].fillna(0.0)
    perf["% Achievement"] = perf.apply(lambda r: r["Total Sales"] / r["Target"] if r["Target"] else 0, axis=1)

    perf = perf.rename(columns={"PMS": "Sales PMS", "AGO": "Sales AGO", "IK": "Sales IK", "Target": "Month Target"})
    perf = perf[["Zone", "Station", "Sales PMS", "Sales AGO", "Sales IK", "Total Sales", "Days Counted", "Average Sales/Day",
                 "Loss PMS", "Loss AGO", "Loss IK", "Total Loss", "% Loss to Sales", "% PMS Loss", "% AGO Loss", "% IK Loss", "Month Target", "% Achievement"]]
    perf = perf.sort_values(["Zone", "Station"])

    network = perf.agg({"Sales PMS": "sum", "Sales AGO": "sum", "Sales IK": "sum", "Total Sales": "sum", "Total Loss": "sum", "Month Target": "sum"}).to_frame().T
    network["% Loss to Sales"] = network["Total Loss"] / network["Total Sales"].replace(0, pd.NA)
    network["% Achievement"] = network["Total Sales"] / network["Month Target"].replace(0, pd.NA)

    banking_summary = banking_df.groupby(["Zone", "Station"], as_index=False).agg(Banking_MTD=("Banking", "sum"), Banking_Days=("Date", "nunique")) if not banking_df.empty else pd.DataFrame(columns=["Zone", "Station", "Banking_MTD", "Banking_Days"])
    if not banking_summary.empty:
        banking_summary["Avg Daily Banking"] = banking_summary["Banking_MTD"] / banking_summary["Banking_Days"].replace(0, pd.NA)
        banking_summary = banking_summary.merge(perf[["Zone", "Station", "Total Sales"]], on=["Zone", "Station"], how="left")
        banking_summary["Banking/Sales Ratio"] = banking_summary["Banking_MTD"] / banking_summary["Total Sales"].replace(0, pd.NA)

    # Intervention rules as agreed:
    # - Sales intervention: station below 85% of monthly target.
    # - Loss intervention: total loss worse than -0.5% of sales, OR PMS/AGO product loss worse than -0.5%.
    #   IK is excluded from product-level loss intervention, but still included in total station loss.
    sales_intervention = perf[perf["% Achievement"] < 0.85].copy().sort_values("% Achievement")
    if not sales_intervention.empty:
        sales_intervention["Reason"] = "Below 85% of monthly target"

    def loss_reason(row):
        reasons = []
        if row["% Loss to Sales"] <= -0.005:
            reasons.append("Total loss above -0.5%")
        if row["% PMS Loss"] <= -0.005:
            reasons.append("PMS loss above -0.5%")
        if row["% AGO Loss"] <= -0.005:
            reasons.append("AGO loss above -0.5%")
        return "; ".join(reasons)

    loss_intervention = perf.copy()
    loss_intervention["Reason"] = loss_intervention.apply(loss_reason, axis=1)
    loss_intervention = loss_intervention[loss_intervention["Reason"] != ""].copy().sort_values("% Loss to Sales")

    return perf, network, banking_summary, sales_intervention, loss_intervention


def style_workbook(wb):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
        for col in ws.columns:
            max_len = 0
            letter = get_column_letter(col[0].column)
            for cell in list(col)[:100]:
                max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 28)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    header = ws.cell(1, cell.column).value
                    if isinstance(header, str) and ("%" in header or "Ratio" in header):
                        cell.number_format = '0.00%'
                    else:
                        cell.number_format = '#,##0.00'
                if isinstance(cell.value, datetime):
                    cell.number_format = 'yyyy-mm-dd'


def add_grand_total_rows(perf: pd.DataFrame, banking_summary: pd.DataFrame):
    """Return display/export tables with a Grand Total row at the bottom.

    The total row is added only after all calculations are complete so it is never
    treated as a station during ranking, intervention or KPI calculations.
    """
    perf_display = perf.copy()
    if not perf_display.empty:
        total = {col: "" for col in perf_display.columns}
        total["Zone"] = ""
        total["Station"] = "GRAND TOTAL"
        for col in ["Sales PMS", "Sales AGO", "Sales IK", "Total Sales", "Days Counted",
                    "Loss PMS", "Loss AGO", "Loss IK", "Total Loss", "Month Target"]:
            if col in perf_display.columns:
                total[col] = perf_display[col].sum()
        total["Average Sales/Day"] = (
            total["Total Sales"] / total["Days Counted"] if total.get("Days Counted", 0) else 0
        )
        total["% Loss to Sales"] = (
            total["Total Loss"] / total["Total Sales"] if total.get("Total Sales", 0) else 0
        )
        total["% PMS Loss"] = (
            total["Loss PMS"] / total["Sales PMS"] if total.get("Sales PMS", 0) else 0
        )
        total["% AGO Loss"] = (
            total["Loss AGO"] / total["Sales AGO"] if total.get("Sales AGO", 0) else 0
        )
        total["% IK Loss"] = (
            total["Loss IK"] / total["Sales IK"] if total.get("Sales IK", 0) else 0
        )
        total["% Achievement"] = (
            total["Total Sales"] / total["Month Target"] if total.get("Month Target", 0) else 0
        )
        perf_display = pd.concat([perf_display, pd.DataFrame([total])], ignore_index=True)

    banking_display = banking_summary.copy()
    if not banking_display.empty:
        total = {col: "" for col in banking_display.columns}
        total["Zone"] = ""
        total["Station"] = "GRAND TOTAL"
        if "Banking_MTD" in banking_display.columns:
            total["Banking_MTD"] = banking_display["Banking_MTD"].sum()
        if "Banking_Days" in banking_display.columns:
            total["Banking_Days"] = banking_display["Banking_Days"].sum()
        if "Total Sales" in banking_display.columns:
            total["Total Sales"] = banking_display["Total Sales"].sum()
        total["Avg Daily Banking"] = (
            total.get("Banking_MTD", 0) / total.get("Banking_Days", 0) if total.get("Banking_Days", 0) else 0
        )
        total["Banking/Sales Ratio"] = (
            total.get("Banking_MTD", 0) / total.get("Total Sales", 0) if total.get("Total Sales", 0) else 0
        )
        banking_display = pd.concat([banking_display, pd.DataFrame([total])], ignore_index=True)

    return perf_display, banking_display



# ----------------------------
# Maintenance Dashboard Module
# ----------------------------
MAINTENANCE_SHEETS = ["MAINTENANCE ISSUES-PRIORITY", "MAINTENANCE ISSUES-LONGTERM"]


def is_pending_repair(value) -> bool:
    """Treat blank, Pending and Not Yet Repaired as open. Treat actual dates/done/repaired/closed as closed."""
    if value is None or clean_text(value) == "":
        return True
    if isinstance(value, (datetime, date)):
        return False
    txt = norm_text(value)
    if any(x in txt for x in ["DONE", "REPAIRED", "CLOSED", "COMPLETE", "COMPLETED"]):
        return False
    if any(x in txt for x in ["PENDING", "NOT YET", "NOT REPAIRED", "ONGOING", "IN PROGRESS"]):
        return True
    return True


def issue_status(value) -> str:
    return "Open" if is_pending_repair(value) else "Closed"


def find_maintenance_sheet_names(wb):
    names = []
    for wanted in MAINTENANCE_SHEETS:
        for s in wb.sheetnames:
            if norm_text(s) == norm_text(wanted) or norm_text(wanted) in norm_text(s):
                if s not in names:
                    names.append(s)
    return names


def extract_maintenance_from_workbook(uploaded_file, zone_name: str) -> pd.DataFrame:
    wb = workbook_from_upload(uploaded_file)
    rows = []
    for sheet_name in find_maintenance_sheet_names(wb):
        ws = wb[sheet_name]
        current_station = ""
        for row in ws.iter_rows(min_row=1, values_only=True):
            vals = list(row) + [None] * 12
            # Standard maintenance templates:
            # B = station, C = issue type/category, D = issue, E = date reported,
            # F = reported to, G = action person, H = date repaired/status, I = comments.
            station_cell = vals[1]
            category = vals[2]
            issue = vals[3]
            date_reported = vals[4]
            reported_to = vals[5]
            action_person = vals[6]
            repaired_status = vals[7]
            comments = vals[8]

            possible_station = norm_station(station_cell)
            if possible_station and not is_total_station(possible_station) and "STATION" not in norm_text(possible_station):
                current_station = possible_station

            if norm_text(issue) in ["BREAKDOWN REPORTED", "DAILY MAINTENANCE REPORT"]:
                continue
            if not current_station:
                continue
            if clean_text(category) == "" and clean_text(issue) == "":
                continue
            # Avoid station heading rows that have no actual issue details.
            if clean_text(issue) == "" and clean_text(date_reported) == "" and clean_text(comments) == "":
                continue

            d = to_date(date_reported)
            status = issue_status(repaired_status)
            if status != "Open":
                continue
            issue_text = clean_text(issue)
            category_text = clean_text(category)
            comment_text = clean_text(comments)
            combined = " ".join([category_text, issue_text, comment_text]).strip()
            if combined == "":
                continue
            rows.append({
                "Zone": zone_name,
                "Station": current_station,
                "Issue Type": category_text,
                "Issue": issue_text,
                "Date Reported": d,
                "Reported To": clean_text(reported_to),
                "Assigned To": clean_text(action_person),
                "Status": "Open",
                "Comments": comment_text,
                "Source Sheet": sheet_name,
            })
    df = pd.DataFrame(rows)
    if df.empty:
        return pd.DataFrame(columns=["Zone", "Station", "Issue Type", "Issue", "Date Reported", "Reported To", "Assigned To", "Status", "Comments", "Source Sheet"])
    return df



def classify_maintenance_issue(issue_type, issue, comments):
    """Classify maintenance using the actual Issue + Comments narrative only.

    The Issue Type column is intentionally ignored because the uploaded templates
    often contain misleading labels such as FENCE, CIVIL WORKS, SALES, etc.
    """
    text = norm_text(" ".join([clean_text(issue), clean_text(comments)]))

    def has_any(words):
        return any(w in text for w in words)

    product_words = ["PMS", "AGO", "IK", "ULP", "UNLEADED", "RMS", "DIESEL", "KEROSENE", "KERO"]
    has_product = has_any(product_words)

    # 1. Automation — card readers, Gold Card, ATG, controllers and automation communication.
    if has_any([
        "CARD READER", "CARD READERS", "GOLD CARD", "VEEDER", "ATG", "AUTOMATION",
        "CONTROLLER", "COMMUNICATION", "POS", "ON MANUAL", "MANUAL CARD", "TAKE CARD"
    ]):
        return "Automation", "Critical", "Automation/control risk"

    # 2. Health & Safety — safety, fire, sand buckets, washrooms/toilets, blocked drainage where it creates HSE exposure.
    if has_any([
        "SAND BUCKET", "SAND BUCKETS", "FIRE EXTINGUISHER", "EXTINGUISHER",
        "SPILL KIT", "SPILL", "EMERGENCY STOP", "E-STOP", "EARTHING",
        "SAFETY", "FIRST AID", "FIRE", "WASHROOM", "WASHROOMS", "TOILET BLOCKED",
        "TOILET CLEANING", "GENTS TOILET", "LADIES TOILET", "NOT CONNECTED TO MAIN SEWER",
        "SEWER LINE", "BEES", "TINTING/CURTAINS", "TINTING", "CURTAINS", "ENTRANCE MAIN DOOR REQUIRE METALLIC GRILLS"
    ]):
        return "Health & Safety", "Critical", "Health and safety exposure"

    # 3. Contamination / product storage water issues.
    if (has_product and has_any(["WATER TRACE", "WATER TRACES", "FULL OF WATER", "CONTAM", "CONTAMINATION", "CONTAMINATED"])) or has_any([
        "FULL OF WATER AFTER HEAVY RAINS", "WATER AFTER HEAVY RAINS"
    ]):
        return "contamination", "Critical", "Product contamination/water ingress"

    # 4. Electrical, Generator & Power — genset/generator, KPLC/Kenya Power, voltage, phase, sockets, lights.
    if has_any([
        "GENSET", "GENERATOR", "ENGINE", "RADIATOR", "BATTERY", "KPLC", "KENYA POWER",
        "POWER", "VOLTAGE", "LOW VOLTAGE", "HIGH VOLTAGE", "PHASE", "3 PHASE", "3PHASE",
        "ELECTRIC", "ELECTRICAL", "SOCKET", "DB BOARD", "CIRCUIT BREAKER", "BREAKER",
        "ATS", "AVR", "STABILIZER", "STABILISER", "INVERTER", "LIGHT", "LIGHTING",
        "SECURITY LIGHT", "UNDERCANOPY LIGHT", "UNDER CANOPY LIGHT", "CANOPY LIGHT",
        "FLOOD LIGHT", "BULB", "SWITCH", "HAND DRIER", "NOT LIGHTING", "DIM"
    ]):
        return "Electrical, Generator & Power", "Critical", "Electrical/generator/power issue"

    # 5. Compressor maintenance.
    if has_any(["COMPRESSOR", "AIR GAUGE", "PRESSURE GAUGE", "PRESSURE NOZZLE", "PRESSURE LINE"]):
        return "compressor maintenance", "Major", "Compressor/air pressure issue"

    # 6. Branding & Appearance — station image, totem, fascia, branding, painting, Petro logo, price display.
    if has_any([
        "PAINT", "PAINTING", "REPAINT", "REPAINTING", "BRANDING", "REBRANDING",
        "TOTEM", "FASCIA", "FACIA", "CLADDING", "SIGN BOARD", "SIGNBOARD", "SIGN BOARDS",
        "SIGNAGE", "PRICE BOARD", "DISPLAYING PRICES", "DISPLAY", "DIRECTIONAL SIGN",
        "PETRO LOGO", "LOGO", "CORPORATE IMAGE", "FACELIFT", "UPPER PART DAMAGED",
        "CRACKED ONE SIDE AT THE BOTTOM", "POWER DISPLAY", "LENS FAULTY", "ROOF SLAP PAINT PEELING"
    ]):
        return "Branding & Appearance", "Major", "Branding/appearance issue"

    # 7. Routine Maintenance & Civil Works — civil/facility problems must be checked before pump logic.
    # This prevents phrases such as 'pump island potholes' from being classified as pump breakdowns.
    if has_any([
        "PERIMETER", "WALL", "FENCE", "POTHOLE", "POTHOLES", "PORTHOLES", "PUMP ISLAND",
        "DRAIN", "DRAINAGE", "CULVERT", "CABRO", "TOILET", "PLUMB", "PLUMBING", "SEPTIC",
        "SEWER", "DOOR", "WINDOW", "WINDOWS", "ROOF", "SLAB", "MASON", "MASONRY",
        "STRUCTURE", "ACCESS", "ENTRANCE", "EXIT", "OFFICE", "WELD", "WELDING",
        "SHADE", "FLOOR", "TILES", "TYPES", "PAVEMENT", "DRIVEWAY", "LEAKING",
        "LEAKS", "HEAVY DOWNPOUR", "RAINY SEASON", "CANOPY LEAK", "CANOPY LEAKING",
        "CANOPY IRON", "CANOPY FACIA", "CANOPY FACIA", "FORECOURT METAL COVERS",
        "FORECOURT DRAINAGE", "MANHOLE", "MAN HOLE", "MANHOLE COVER", "MANHOLE COVERS",
        "PRODUCT MANHOLES", "HINGES", "HIDGES", "COVER BROKEN", "COVER RAILS", "LOUVRES",
        "WINDOW PANES", "CONCRETE", "SCREEDING", "FUELING RAMP", "FUELLING RAMP",
        "PUMP GUARD", "PUMP GUARDS", "BOLLARD", "BOLLARDS", "MOUNT HOOKS", "MOUNTING AT THE FORECOURT"
    ]):
        return "Routine Maintenance & Civil Works", "Major", "Routine maintenance/civil works issue"

    # 8. Pumps & Dispensers — only actual pump/dispenser breakdowns, not civil works near pumps.
    if has_any([
        "PUMP FAULT", "PUMP FAULTY", "PUMP INSTALLED", "PUMP NOZZLE", "PUMP NOZZLES",
        "SUBMERSIBLE PUMP", "SUBMASSIBLE PUMP", "UNABLE TO PUMP", "UNABLE TO DISPENSE",
        "NOT DISPENSING", "DISPENSING BUT NOT COUNTING", "DISPENSING SLOWLY", "DISPENSE",
        "DISPENSING", "NOZZLE", "NOZZLES", "KEYPAD", "METERING UNIT", "METERING",
        "SOLENOID", "UNDERVALVE", "HOSE", "HOSE LEAKING", "HOSE REPLACEMENT",
        "MOTOR FAULTY", "NOZZLE BOOT", "OVERDISPENSE", "PMS1", "PMS2", "PMS3", "PMS4",
        "AGO1", "AGO2", "AGO3", "AGO4", "AGO5", "AGO6", "IK PUMP", "ERROR 20", "ERRO 20",
        "POWER BOARD FAULTY", "MOTHER BOARD", "BOARD TRANSFERRED", "RUNNING CONTINOUS", "RUNNING CONTINUOUS",
        "AIRLOCK", "AIRLOCK REMOVAL", "AFTER PRICE CHANGE", "AFTER PICE CHANGE"
    ]):
        return "Pumps & Dispensers", "Critical", "Pump/dispenser breakdown"

    # 9. Other.
    return "Other", "Major", "Pending maintenance issue"


def build_maintenance_reports(maintenance_df: pd.DataFrame):
    if maintenance_df.empty:
        empty = pd.DataFrame()
        return empty, empty, empty, empty, empty

    today = date.today()
    df = maintenance_df.copy()

    # Support both old and new extracted column names.
    if "Reported By" not in df.columns:
        if "Reported To" in df.columns:
            df["Reported By"] = df["Reported To"]
        else:
            df["Reported By"] = ""
    if "Action Person" not in df.columns:
        df["Action Person"] = ""
    if "Reported To" not in df.columns:
        if "Assigned To" in df.columns:
            df["Reported To"] = df["Assigned To"]
        else:
            df["Reported To"] = ""

    df["Date Reported"] = df["Date Reported"].apply(to_date)

    def calculate_days_open(d):
        if d is None or pd.isna(d):
            return None
        return (today - d).days

    df["Days Open"] = df["Date Reported"].apply(calculate_days_open)

    # IMPORTANT: classifier ignores Issue Type and uses Issue + Comments only.
    classifications = df.apply(
        lambda r: classify_maintenance_issue(None, r.get("Issue"), r.get("Comments")),
        axis=1,
    )
    df["Category"] = [x[0] for x in classifications]
    df["Priority"] = [x[1] for x in classifications]
    df["Management Reason"] = [x[2] for x in classifications]

    action_person_map = {
        "Pumps & Dispensers": "Vijay",
        "Automation": "Ochol/ Steve",
        "Branding & Appearance": "Patel/ Vijay",
        "compressor maintenance": "Vijay",
        "Compressor Maintenance": "Vijay",
        "contamination": "Vijay",
        "Electrical, Generator & Power": "Vijay",
        "Health & Safety": "Edward",
        "Routine Maintenance & Civil Works": "Milan",
        "Other": "Other",
    }

    category_order = {
        "Pumps & Dispensers": 1,
        "Automation": 2,
        "Electrical, Generator & Power": 3,
        "contamination": 4,
        "compressor maintenance": 5,
        "Branding & Appearance": 6,
        "Health & Safety": 7,
        "Routine Maintenance & Civil Works": 8,
        "Other": 9,
    }
    priority_order = {"Critical": 1, "Major": 2, "Minor": 3}

    df["Action Person"] = df["Category"].map(action_person_map).fillna("Other")
    df["Reported By"] = df["Reported By"].apply(clean_text)
    df["Reported To"] = df["Reported To"].apply(clean_text).replace("", "Vijay")
    df["Issue ID"] = [f"MNT-{today.strftime('%Y%m%d')}-{i+1:03d}" for i in range(len(df))]
    df["_Category Order"] = df["Category"].map(category_order).fillna(99)
    df["_Priority Order"] = df["Priority"].map(priority_order).fillna(99)

    register_columns = [
        "Issue ID", "Zone", "Station", "Category", "Action Person", "Priority",
        "Issue Type", "Issue", "Date Reported", "Days Open",
        "Reported By", "Reported To", "Management Reason", "Comments", "Source Sheet",
        "_Category Order", "_Priority Order",
    ]
    for col in register_columns:
        if col not in df.columns:
            df[col] = ""

    register = df[register_columns].copy()
    register = register.sort_values(
        ["_Category Order", "_Priority Order", "Days Open", "Zone", "Station"],
        ascending=[True, True, False, True, True],
    ).drop(columns=["_Category Order", "_Priority Order"])

    station_summary = register.groupby(["Zone", "Station"], as_index=False).agg(
        Open_Issues=("Issue ID", "count"),
        Critical_Issues=("Priority", lambda s: (s == "Critical").sum()),
        Oldest_Days_Open=("Days Open", "max"),
    ).sort_values(["Critical_Issues", "Open_Issues", "Oldest_Days_Open"], ascending=[False, False, False])

    aging = register.copy()

    def age_bucket(x):
        if pd.isna(x):
            return "Unknown Date"
        if x <= 30:
            return "0-30 Days"
        if x <= 90:
            return "31-90 Days"
        if x <= 180:
            return "91-180 Days"
        return "Above 180 Days"

    aging["Age Bucket"] = aging["Days Open"].apply(age_bucket)
    aging_report = aging.groupby("Age Bucket", as_index=False).agg(Issues=("Issue ID", "count"))

    critical = register[(register["Priority"] == "Critical") | (register["Days Open"].fillna(0) > 90)].copy()
    critical["_Category Order"] = critical["Category"].map(category_order).fillna(99)
    critical["_Priority Order"] = critical["Priority"].map(priority_order).fillna(99)
    critical = critical.sort_values(
        ["_Category Order", "_Priority Order", "Days Open", "Zone", "Station"],
        ascending=[True, True, False, True, True],
    ).drop(columns=["_Category Order", "_Priority Order"])

    responsibility = register.groupby("Action Person", as_index=False).agg(
        Open_Issues=("Issue ID", "count"),
        Critical_Issues=("Priority", lambda s: (s == "Critical").sum()),
        Stations_Affected=("Station", "nunique"),
        Oldest_Days_Open=("Days Open", "max"),
    ).sort_values(["Critical_Issues", "Open_Issues"], ascending=[False, False])

    return register, station_summary, aging_report, critical, responsibility

def to_maintenance_excel_bytes(maintenance_reports):
    register, station_summary, aging_report, critical, responsibility = maintenance_reports
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        register.to_excel(writer, sheet_name="Pending Register", index=False)
        station_summary.to_excel(writer, sheet_name="Station Summary", index=False)
        aging_report.to_excel(writer, sheet_name="Aging Analysis", index=False)
        critical.to_excel(writer, sheet_name="Critical Issues", index=False)
        responsibility.to_excel(writer, sheet_name="Action Person Summary", index=False)
        style_workbook(writer.book)
    return output.getvalue()


def to_excel_bytes(reports):
    perf, network, banking_summary, sales_intervention, loss_intervention = reports
    perf_display, banking_display = add_grand_total_rows(perf, banking_summary)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        network.to_excel(writer, sheet_name="Executive Summary", index=False)
        perf_display.to_excel(writer, sheet_name="MTD Performance", index=False)
        banking_display.to_excel(writer, sheet_name="Banking", index=False)
        sales_intervention.to_excel(writer, sheet_name="Sales Intervention", index=False)
        loss_intervention.to_excel(writer, sheet_name="Loss Intervention", index=False)
        style_workbook(writer.book)
        for ws_name in ["MTD Performance", "Banking"]:
            ws = writer.book[ws_name]
            last_row = ws.max_row
            for cell in ws[last_row]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9EAD3")
    return output.getvalue()



def format_for_display(df: pd.DataFrame) -> pd.DataFrame:
    """Format percentages for Streamlit display without changing calculation tables."""
    out = df.copy()
    percent_cols = [c for c in out.columns if ("%" in str(c) or "Ratio" in str(c))]
    for c in percent_cols:
        out[c] = out[c].apply(lambda x: "" if pd.isna(x) or x == "" else f"{float(x):.2%}")
    return out


st.set_page_config(page_title="Petro Oil Management Report Generator", layout="wide")
st.title("Petro Oil Management Report Generator")
st.caption("Upload daily sales and maintenance workbooks to generate management reports and intervention dashboards.")

with st.sidebar:
    st.header("Upload Daily Zone Files")
    coast_file = st.file_uploader("Coast Daily Sales Workbook", type=["xlsx"], key="coast")
    nrb_file = st.file_uploader("Nairobi Daily Sales Workbook", type=["xlsx"], key="nrb")
    western_file = st.file_uploader("Western Daily Sales Workbook", type=["xlsx"], key="western")

st.sidebar.divider()
st.sidebar.header("Maintenance Files")
st.sidebar.caption("Upload regional daily maintenance reports to generate one pending maintenance intervention report.")
maint_coast_file = st.sidebar.file_uploader("Coast Maintenance Workbook", type=["xlsx"], key="maint_coast")
maint_nrb_file = st.sidebar.file_uploader("Nairobi/Mt Kenya Maintenance Workbook", type=["xlsx"], key="maint_nrb")
maint_western_file = st.sidebar.file_uploader("Western Maintenance Workbook", type=["xlsx"], key="maint_western")

st.sidebar.divider()
st.sidebar.header("Monthly Targets")
st.sidebar.caption("Targets are saved in the app after the first upload and reused in future reports.")
target_file = st.sidebar.file_uploader("Upload/Replace Monthly Target File", type=["xlsx"], key="target")

if target_file is not None and st.sidebar.button("Save Uploaded Targets"):
    uploaded_targets = extract_targets(target_file)
    if uploaded_targets.empty:
        st.sidebar.error("No targets found. File must contain Station and Month Target/Target columns.")
    else:
        saved = save_targets(uploaded_targets)
        st.session_state["targets_df"] = saved
        st.sidebar.success(f"Saved {len(saved)} station targets.")

if "targets_df" not in st.session_state:
    st.session_state["targets_df"] = load_saved_targets()

with st.sidebar.expander("View/Edit Saved Targets", expanded=False):
    edited_targets = st.data_editor(
        st.session_state["targets_df"],
        num_rows="dynamic",
        use_container_width=True,
        hide_index=True,
        column_config={"Target": st.column_config.NumberColumn("Target", format="%.0f")},
        key="targets_editor",
    )
    if st.button("Save Edited Targets"):
        saved = save_targets(edited_targets)
        st.session_state["targets_df"] = saved
        st.success(f"Targets updated: {len(saved)} stations saved.")

st.sidebar.caption(f"Saved targets: {len(st.session_state['targets_df'])} stations")

st.sidebar.divider()
st.sidebar.header("Station Master")
st.sidebar.caption("Approved station names and aliases. The app maps only what appears here; anything else is sent to Unmapped Stations for review.")
if "station_master_df" not in st.session_state:
    st.session_state["station_master_df"] = load_station_master()

with st.sidebar.expander("View/Edit Station Master", expanded=False):
    edited_station_master = st.data_editor(
        st.session_state["station_master_df"],
        num_rows="dynamic",
        use_container_width=True,
        hide_index=True,
        key="station_master_editor",
    )
    if st.button("Save Station Master"):
        saved_master = save_station_master(edited_station_master)
        st.session_state["station_master_df"] = saved_master
        st.success(f"Station Master updated: {len(saved_master)} approved station/alias rows saved.")

st.sidebar.caption(f"Station master rows: {len(st.session_state['station_master_df'])}")

uploaded = [("COAST", coast_file), ("NAIROBI", nrb_file), ("WESTERN", western_file)]
maintenance_uploaded = [("COAST", maint_coast_file), ("NAIROBI", maint_nrb_file), ("WESTERN", maint_western_file)]
ready = any(f is not None for _, f in uploaded)
maintenance_ready = any(f is not None for _, f in maintenance_uploaded)

if st.session_state.get("targets_df", pd.DataFrame()).empty:
    st.warning("No monthly targets are saved yet. Upload the target file once in the sidebar, then click Save Uploaded Targets. Reports can still run, but achievement will show as 0 until targets are saved.")

if not ready and not maintenance_ready:
    st.info("Upload at least one sales workbook or one maintenance workbook to begin.")
    st.stop()

if ready and st.button("Generate Sales/Loss/Banking Reports", type="primary"):
    sales_frames, loss_frames, banking_frames = [], [], []
    errors = []
    for zone, f in uploaded:
        if f is None:
            continue
        try:
            wb = workbook_from_upload(f)
            sales_frames.append(extract_product_summary(wb, zone))
            loss_frames.append(extract_product_loss(wb, zone))
            banking_frames.append(extract_banking(wb, zone))
        except Exception as e:
            errors.append(f"{zone}: {e}")

    if errors:
        st.error("Some files could not be processed:\n" + "\n".join(errors))
    sales_df = pd.concat(sales_frames, ignore_index=True) if sales_frames else pd.DataFrame()
    loss_df = pd.concat(loss_frames, ignore_index=True) if loss_frames else pd.DataFrame()
    banking_df = pd.concat(banking_frames, ignore_index=True) if banking_frames else pd.DataFrame()
    targets_df = st.session_state.get("targets_df", load_saved_targets())
    station_master_df = st.session_state.get("station_master_df", load_station_master())

    sales_df, unmapped_sales = apply_station_master_to_df(sales_df, station_master_df, "Sales")
    loss_df, unmapped_losses = apply_station_master_to_df(loss_df, station_master_df, "Loss")
    banking_df, unmapped_banking = apply_station_master_to_df(banking_df, station_master_df, "Banking")
    unmapped_df = pd.concat([unmapped_sales, unmapped_losses, unmapped_banking], ignore_index=True)
    st.session_state["unmapped_stations"] = unmapped_df

    if not unmapped_df.empty:
        st.warning("Some station names were not in the approved Station Master and were excluded from the report. Open the Unmapped Stations tab, add aliases to Station Master, save, then regenerate.")

    reports = build_reports(sales_df, loss_df, banking_df, targets_df)
    perf, network, banking_summary, sales_intervention, loss_intervention = reports

    st.session_state["reports"] = reports
    st.session_state["raw"] = (sales_df, loss_df, banking_df)


if maintenance_ready and st.button("Generate Maintenance Dashboard", type="secondary"):
    maintenance_frames = []
    maintenance_errors = []
    for zone, f in maintenance_uploaded:
        if f is None:
            continue
        try:
            maintenance_frames.append(extract_maintenance_from_workbook(f, zone))
        except Exception as e:
            maintenance_errors.append(f"{zone}: {e}")
    if maintenance_errors:
        st.error("Some maintenance files could not be processed:\n" + "\n".join(maintenance_errors))
    maintenance_df = pd.concat(maintenance_frames, ignore_index=True) if maintenance_frames else pd.DataFrame()
    station_master_df = st.session_state.get("station_master_df", load_station_master())
    maintenance_df, unmapped_maintenance = apply_station_master_to_df(maintenance_df, station_master_df, "Maintenance")
    st.session_state["unmapped_maintenance"] = unmapped_maintenance
    if not unmapped_maintenance.empty:
        st.warning("Some maintenance station names were not in Station Master and were excluded. Add aliases, save Station Master, then regenerate.")
    maintenance_reports = build_maintenance_reports(maintenance_df)
    st.session_state["maintenance_reports"] = maintenance_reports
    st.session_state["maintenance_raw"] = maintenance_df

if "maintenance_reports" in st.session_state:
    register, station_summary, aging_report, critical, responsibility = st.session_state["maintenance_reports"]
    st.header("Maintenance Intervention Dashboard")
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Open Issues", f"{len(register):,.0f}")
    m2.metric("Critical Issues", f"{(register['Priority'] == 'Critical').sum():,.0f}" if not register.empty else "0")
    m3.metric("Stations Affected", f"{register['Station'].nunique():,.0f}" if not register.empty else "0")
    m4.metric("Issues >90 Days", f"{(register['Days Open'].fillna(0) > 90).sum():,.0f}" if not register.empty else "0")

    mtabs = st.tabs(["Pending Register", "Critical Issues", "Station Summary", "Aging Analysis", "Action Person Summary", "Unmapped Maintenance"])
    with mtabs[0]:
        st.caption("Station-specific pending maintenance register. Original issue wording is preserved; categories are added only for management prioritization.")
        st.dataframe(register, use_container_width=True, hide_index=True)
    with mtabs[1]:
        st.caption("Includes sales-impact issues, product storage risks, product immobilization, automation risks and any issue older than 90 days.")
        st.dataframe(critical, use_container_width=True, hide_index=True)
    with mtabs[2]:
        st.dataframe(station_summary, use_container_width=True, hide_index=True)
    with mtabs[3]:
        st.dataframe(aging_report, use_container_width=True, hide_index=True)
    with mtabs[4]:
        st.dataframe(responsibility, use_container_width=True, hide_index=True)
    with mtabs[5]:
        unmapped_maintenance = st.session_state.get("unmapped_maintenance", pd.DataFrame())
        st.caption("These were excluded because they are not approved in Station Master. Add them as aliases and regenerate.")
        st.dataframe(unmapped_maintenance, use_container_width=True, hide_index=True)

    maintenance_excel = to_maintenance_excel_bytes(st.session_state["maintenance_reports"])
    st.download_button("Download Excel Maintenance Report", data=maintenance_excel, file_name="petro_pending_maintenance_report.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if "reports" in st.session_state:
    perf, network, banking_summary, sales_intervention, loss_intervention = st.session_state["reports"]
    perf_display, banking_display = add_grand_total_rows(perf, banking_summary)
    k = network.iloc[0]
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Network Sales", f"{k['Total Sales']:,.0f} L")
    c2.metric("PMS", f"{k['Sales PMS']:,.0f} L")
    c3.metric("AGO", f"{k['Sales AGO']:,.0f} L")
    c4.metric("IK", f"{k['Sales IK']:,.0f} L")
    c5.metric("Achievement", f"{k['% Achievement']:.1%}" if pd.notna(k['% Achievement']) else "N/A")

    tabs = st.tabs(["MTD Performance", "Banking", "Sales Intervention", "Loss Intervention", "Unmapped Stations", "Raw Data"])
    with tabs[0]:
        st.dataframe(format_for_display(perf_display), use_container_width=True, hide_index=True)
    with tabs[1]:
        st.dataframe(format_for_display(banking_display), use_container_width=True, hide_index=True)
    with tabs[2]:
        st.caption("Rule: station appears here when achievement is below 85% of monthly target.")
        st.dataframe(format_for_display(sales_intervention), use_container_width=True, hide_index=True)
    with tabs[3]:
        st.caption("Rule: station appears here when total loss is worse than -0.5%, or PMS/AGO product loss is worse than -0.5%. IK is excluded from product-level intervention.")
        st.dataframe(format_for_display(loss_intervention), use_container_width=True, hide_index=True)
    with tabs[4]:
        st.subheader("Unmapped Station Names")
        st.caption("These were excluded because they are not approved in Station Master. Add them as aliases, save Station Master, then regenerate reports.")
        unmapped_df = st.session_state.get("unmapped_stations", pd.DataFrame())
        st.dataframe(unmapped_df, use_container_width=True, hide_index=True)

    with tabs[5]:
        sales_df, loss_df, banking_df = st.session_state["raw"]
        st.subheader("Extracted Sales")
        st.dataframe(sales_df.head(500), use_container_width=True, hide_index=True)
        st.subheader("Extracted Losses")
        st.dataframe(loss_df.head(500), use_container_width=True, hide_index=True)
        st.subheader("Extracted Banking")
        st.dataframe(banking_df.head(500), use_container_width=True, hide_index=True)

    excel_bytes = to_excel_bytes(st.session_state["reports"])
    st.download_button("Download Excel Management Report", data=excel_bytes, file_name="petro_management_report.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
